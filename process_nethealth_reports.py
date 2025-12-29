#!/usr/bin/env python3
"""
NetHealth Report Processor
Reads 4 Excel reports from NetHealth and generates facility_data.json for TheraScope
"""

import pandas as pd
import json
from datetime import datetime
import sys
import re

# Define regions
GOLDEN_COAST = [
    'Los Altos Post Acute',
    'Mountain View HC',
    'Camino Ridge Post Acute',
    'The Win Post Acute',
    'PAC Hills Post Acute',
    'Morgan Hill HC',
    'Manresa HC',
    'Gilroy HC',
    'Pac Coast PA'
]

OVERLAND = [
    'Palo Alto Post Acute',
    'Belmont HC',
    'Eden HC',
    'Golden Harbor HC',
    'West Shore PA',
    'Capital PA',
    'Bridgewood PA',
    'Cedarwood PA'
]

# Default Med B Eligible counts (these don't change much week-to-week)
# Update these manually if you know they've changed significantly
MED_B_ELIGIBLE_DEFAULTS = {
    'Belmont HC': 8,
    'Belmont Healthcare Center': 8,
    'Bridgewood PA': 15,
    'Camino Ridge Post Acute': 18,
    'Capital PA': 26,
    'Cedarwood PA': 15,
    'Eden HC': 68,
    'Eden Healthcare Center': 68,
    'Gilroy HC': 40,
    'Golden Harbor HC': 21,
    'Golden Harbor Healthcare Center': 21,
    'Los Altos Post Acute': 27,
    'Los Altos Sub-Acute': 27,
    'Manresa HC': 25,
    'Morgan Hill HC': 16,
    'Morgan Hill Healthcare Center': 16,
    'Mountain View HC': 31,
    'Mountain View': 31,
    'PAC Hills Post Acute': 17,
    'Pacific Hills Manor': 17,
    'Pac Coast PA': 35,
    'Palo Alto Post Acute': 6,
    'Palo Alto Sub-Acute & Rehab Center': 6,
    'The Win Post Acute': 50,
    'Mission Skilled Nursing & SubA': 50,  # Same as The Win
    'West Shore PA': 43,
}

def get_region(facility_name):
    """Determine which region a facility belongs to"""
    normalized = facility_name.strip()
    
    for gc_facility in GOLDEN_COAST:
        if gc_facility.lower() in normalized.lower() or normalized.lower() in gc_facility.lower():
            return 'Golden Coast'
    
    for ov_facility in OVERLAND:
        if ov_facility.lower() in normalized.lower() or normalized.lower() in ov_facility.lower():
            return 'Overland'
    
    return 'Unknown'

def normalize_facility_name(name):
    """Normalize facility names to match our standard format"""
    name = str(name).strip()
    
    # Remove prefixes like *PWBK, PWB, etc.
    name = re.sub(r'^\*?PWBK\s+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'^\*?PWB\s+', '', name, flags=re.IGNORECASE)
    
    # Specific mappings
    mapping = {
        'Gilroy Healthcare & Rehab Cent': 'Gilroy HC',
        'Grant Cuesta Nursing & Rehab H': 'Camino Ridge Post Acute',  # Same as Camino Ridge
        'Morgan Hill Healthcare Center': 'Morgan Hill HC',
        'Mountain View Healthcare Cente': 'Mountain View HC',
        'Mountain View': 'Mountain View HC',
        'Pacific Coast Manor': 'Pac Coast PA',
        'Los Altos Sub-Acute': 'Los Altos Post Acute',
        'Los Altos Sub-Acute & Rehab Ce': 'Los Altos Post Acute',
        'Palo Alto Sub-Acute & Rehab Center': 'Palo Alto Post Acute',
        'Palo Alto Sub Acute': 'Palo Alto Post Acute',
        'Mission Skilled Nursing & SubA': 'The Win Post Acute',  # Same as The Win
        'Eden Healthcare Center': 'Eden HC',
        'Eden Post Acute Care': 'Eden HC',
        'Belmont Healthcare Center': 'Belmont HC',
        'Belmont Hills Health & Rehab': 'Belmont HC',
        'Golden Harbor Healthcare Center': 'Golden Harbor HC',
        'Golden Harbor Nursing & Rehab': 'Golden Harbor HC',
        'West Shore Post Acute': 'West Shore PA',
        'Capital Post Acute': 'Capital PA',
        'Bridgewood Post Acute': 'Bridgewood PA',
        'Cedarwood Post Acute': 'Cedarwood PA',
        'Pacific Hills Manor': 'PAC Hills Post Acute',
        'Manresa Healthcare Center': 'Manresa HC'
    }
    
    # Check for exact match first (case-insensitive)
    for old, new in mapping.items():
        if name.lower() == old.lower():
            return new
    
    return name

def extract_week_from_date_range(date_range_str):
    """Extract week identifier from date range like '12/14/2025 - 12/20/2025'"""
    try:
        # Extract start date
        start_date = date_range_str.split('-')[0].strip()
        # Parse it
        dt = datetime.strptime(start_date, '%m/%d/%Y')
        # Return as MMDD format
        return dt.strftime('%m%d').lstrip('0')
    except:
        return None

def process_productivity_report(filepath):
    """Extract productivity data by facility from NetHealth productivity report"""
    xl = pd.ExcelFile(filepath)
    productivity_data = {}
    
    # Read each sheet (each sheet is a facility)
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        
        current_facility = None
        last_totals_productivity = None
        
        for idx, row in df.iterrows():
            # Look for "Site:" lines
            if pd.notna(row.get(1)) and 'Site:' in str(row.get(1)):
                facility_name = str(row[1]).replace('Site:', '').strip()
                current_facility = normalize_facility_name(facility_name)
            
            # Look for ALL TOTALS rows and keep track of the last one
            if pd.notna(row.get(1)) and 'TOTALS' in str(row.get(1)) and current_facility:
                # Efficiency is in column 14 (last column) - this is what we want
                if len(row) > 14:
                    productivity = row.get(14)
                    if pd.notna(productivity):
                        # Remove % and convert to float
                        if isinstance(productivity, str):
                            last_totals_productivity = float(productivity.replace('%', ''))
                        else:
                            last_totals_productivity = float(productivity) * 100
        
        # Save the last TOTALS productivity found (which is the grand total)
        if current_facility and last_totals_productivity is not None:
            productivity_data[current_facility] = round(last_totals_productivity, 1)
    
    return productivity_data

def process_cpm_report(filepath):
    """Extract CPM (Expense Per Tx Minute) by facility from NetHealth CPM report"""
    xl = pd.ExcelFile(filepath)
    cpm_data = {}
    
    # Read each sheet (each sheet is a facility)
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        
        current_facility = None
        
        for idx, row in df.iterrows():
            # Look for "Site of Service" lines
            if pd.notna(row.get(0)) and 'Site of Service' in str(row.get(0)):
                if len(row) > 4 and pd.notna(row.get(4)):
                    facility_name = str(row[4]).strip()
                    current_facility = normalize_facility_name(facility_name)
            
            # Look for TOTAL EXPENSES row
            if pd.notna(row.get(0)) and 'TOTAL EXPENSES' in str(row.get(0)) and current_facility:
                # Expense Per TX Minute is in column 12
                if len(row) > 12 and pd.notna(row.get(12)):
                    cpm = row[12]
                    # Remove $ and convert to float
                    if isinstance(cpm, str):
                        cpm = float(cpm.replace('$', '').replace(',', '').replace('(', '').replace(')', ''))
                    else:
                        cpm = abs(float(cpm))
                    
                    cpm_data[current_facility] = round(cpm, 2)
                    break  # Found it for this sheet, move to next
    
    return cpm_data

def process_census_report(filepath):
    """Extract Medicare Part B patient counts by facility from census report"""
    xl = pd.ExcelFile(filepath)
    census_data = {}
    
    # Read each sheet (each sheet is a facility)
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        
        # Find the facility name from row 6
        facility_row = df[df[0] == 'Site(s) of Service:']
        if len(facility_row) > 0:
            facility_name = facility_row.iloc[0, 1]
            facility_name = normalize_facility_name(facility_name)
        else:
            continue
        
        # Count patients - they start after the header row
        # Look for rows where column 1 is 'Medicare Part B'
        patient_rows = df[df[1] == 'Medicare Part B']
        med_b_count = len(patient_rows)
        
        census_data[facility_name] = med_b_count
    
    return census_data

def process_mode_of_treatment_report(filepath):
    """Extract Mode of Treatment (C/G %) by facility from mode of treatment report"""
    df = pd.read_excel(filepath, header=None)
    
    mode_data = {}
    
    # Find the header row (contains "Facility")
    header_idx = None
    for idx, row in df.iterrows():
        if pd.notna(row[0]) and 'Facility' in str(row[0]):
            header_idx = idx
            break
    
    if header_idx is None:
        return mode_data
    
    # Data rows start after header
    for idx in range(header_idx + 1, len(df)):
        row = df.iloc[idx]
        
        # Facility name is in column 0
        facility_name = row[0]
        if pd.isna(facility_name) or facility_name == '':
            continue
        
        facility_name = normalize_facility_name(facility_name)
        
        # C/G % TOTAL is in column 30 (AE) - the last C/G % column
        if len(row) > 30:
            mode_pct = row[30]
            if pd.notna(mode_pct):
                # Convert from decimal to percentage
                mode_pct = float(mode_pct) * 100
                mode_data[facility_name] = round(mode_pct, 2)
    
    return mode_data

def process_units_per_visit_report(filepath):
    """Extract Units Per Visit by facility from units per visit report"""
    df = pd.read_excel(filepath, header=None)
    
    upv_data = {}
    
    # Find all "Total for [Facility]" rows
    for idx, row in df.iterrows():
        if pd.notna(row.get(1)) and 'Total for' in str(row.get(1)):
            facility_raw = str(row.get(1)).replace('Total for ', '').strip()
            facility_name = normalize_facility_name(facility_raw)
            
            # Units Per Visit is in column 8
            if len(row) > 8 and pd.notna(row.get(8)):
                upv = row[8]
                # It's already a number, just convert to float
                if isinstance(upv, str):
                    upv = float(upv)
                upv_data[facility_name] = round(float(upv), 2)
    
    return upv_data

def process_med_b_units_report(filepath):
    """Extract Med B Units billed from multi-sheet report"""
    import openpyxl
    
    wb = openpyxl.load_workbook(filepath)
    units_data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Find facility name (Row 2, Column B)
        facility_raw = ws['B2'].value
        if not facility_raw:
            continue
        
        facility_name = normalize_facility_name(facility_raw)
        
        # Find Total row (last row with "Total" in column A)
        for row in ws.iter_rows():
            if row[0].value == 'Total':
                # Total Units is in column J (index 9)
                total_units = row[9].value
                if total_units is not None:
                    units_data[facility_name] = int(total_units)
                break
    
    return units_data


def process_all_reports(productivity_file, cpm_file, census_file, mode_file, upv_file, medb_units_file):
    """Process all 6 NetHealth reports and generate facility_data.json"""
    
    print("Processing NetHealth Reports...")
    print("=" * 80)
    
    # Extract week from productivity report
    prod_df = pd.read_excel(productivity_file, header=None)
    date_range_row = prod_df[prod_df[1] == 'Date Range:']
    if len(date_range_row) > 0:
        date_range = date_range_row.iloc[0, 2]
        week = extract_week_from_date_range(date_range)
        # Convert to date format
        start_date = date_range.split('-')[0].strip()
        date_obj = datetime.strptime(start_date, '%m/%d/%Y')
        date_str = date_obj.strftime('%Y-%m-%d')
    else:
        week = '1215'  # Default
        date_str = '2025-12-15'
    
    print(f"Week: {week}")
    print(f"Date: {date_str}")
    print()
    
    # Process each report
    print("1. Processing Productivity Report...")
    productivity_data = process_productivity_report(productivity_file)
    print(f"   Found {len(productivity_data)} facilities")
    
    print("2. Processing CPM Report...")
    cpm_data = process_cpm_report(cpm_file)
    print(f"   Found {len(cpm_data)} facilities")
    
    print("3. Processing Census Report...")
    census_data = process_census_report(census_file)
    print(f"   Found {len(census_data)} facilities with Med B data")
    
    print("4. Processing Mode of Treatment Report...")
    mode_data = process_mode_of_treatment_report(mode_file)
    print(f"   Found {len(mode_data)} facilities")
    
    print("5. Processing Units Per Visit Report...")
    upv_data = process_units_per_visit_report(upv_file)
    print(f"   Found {len(upv_data)} facilities")
    
    print("6. Processing Med B Units Report...")
    medb_units_data = process_med_b_units_report(medb_units_file)
    print(f"   Found {len(medb_units_data)} facilities")
    
    print()
    print("Combining data...")
    print("-" * 80)
    
    # Combine all data
    all_facilities = set()
    all_facilities.update(productivity_data.keys())
    all_facilities.update(cpm_data.keys())
    all_facilities.update(census_data.keys())
    all_facilities.update(mode_data.keys())
    
    records = []
    for facility in sorted(all_facilities):
        if facility == 'Unknown' or facility == '' or facility == 'Total':
            continue
        
        record = {
            'week': week,
            'facility': facility,
            'region': get_region(facility),
            'productivity': productivity_data.get(facility, 0),
            'cpm': cpm_data.get(facility, 0),
            'medBEligible': MED_B_ELIGIBLE_DEFAULTS.get(facility, 0),
            'medBCaseload': census_data.get(facility, 0),
            'medBUnitsThisWeek': medb_units_data.get(facility, 0),
            'unitsPerVisit': upv_data.get(facility, 0),
            'date': date_str
        }
        
        # Add mode of treatment if available
        if facility in mode_data:
            record['modeOfTreatment'] = mode_data[facility]
        
        records.append(record)
        print(f"✓ {facility:40} | Prod: {record['productivity']:5.1f}% | CPM: ${record['cpm']:5.2f} | Med B: {record['medBCaseload']:2}/{record['medBEligible']:2} | Units: {record['medBUnitsThisWeek']:3} | UPV: {record['unitsPerVisit']:.2f}")
    
    print()
    print(f"Total facilities processed: {len(records)}")
    
    return records

if __name__ == '__main__':
    if len(sys.argv) != 7:
        print("Error: Please provide all 6 required NetHealth reports")
        print()
        print("Usage: python process_nethealth_reports.py <productivity.xlsx> <cpm.xlsx> <census.xlsx> <mode.xlsx> <units_per_visit.xlsx> <med_b_units.xlsx>")
        print()
        print("Example:")
        print("  python process_nethealth_reports.py prod.xlsx cpm.xlsx census.xlsx mode.xlsx upv.xlsx medb.xlsx")
        sys.exit(1)
    
    productivity_file = sys.argv[1]
    cpm_file = sys.argv[2]
    census_file = sys.argv[3]
    mode_file = sys.argv[4]
    upv_file = sys.argv[5]
    medb_units_file = sys.argv[6]
    
    records = process_all_reports(productivity_file, cpm_file, census_file, mode_file, upv_file, medb_units_file)
    
    # Save to JSON
    output_file = 'facility_data_nethealth.json'
    with open(output_file, 'w') as f:
        json.dump(records, f, indent=2)
    
    print()
    print(f"✅ Output saved to: {output_file}")
    print()
    print("Next steps:")
    print("1. Review the output file")
    print("2. If it looks good, append this data to your existing facility_data.json")
    print("3. Upload to GitHub!")
